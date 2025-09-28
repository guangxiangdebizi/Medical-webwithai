from typing import Any, Dict, List, Optional
from pathlib import Path
from urllib.parse import urlparse
import os
import csv
import zipfile

from langchain_core.tools import StructuredTool
from pydantic import BaseModel, Field


def create_basic_tools() -> List[StructuredTool]:
    """Create basic non-domain tools (e.g., file preview for user uploads)."""

    class PreviewUploadedFileArgs(BaseModel):
        url: str = Field(description="Uploaded file URL or path starting with /uploads/... or full http(s) URL")
        sheet: Optional[str] = Field(default=None, description="Sheet name for Excel (.xlsx); default first sheet")
        head: int = Field(default=20, description="Number of preview rows to return (default 20)")

    def _resolve_upload_path(url_or_path: str) -> Path:
        """Map a /uploads/... url or full http(s) URL to local filesystem path under backend/uploads."""
        p = str(url_or_path or "").strip()
        if not p:
            raise ValueError("Empty url/path")
        try:
            parsed = urlparse(p)
            if parsed.scheme in ("http", "https"):
                p = parsed.path or p
        except Exception:
            pass
        if "/uploads/" not in p:
            raise ValueError("Only files under /uploads are accessible")
        rel = p.split("/uploads/", 1)[1]
        base = Path(__file__).parent / "uploads"
        full = (base / rel).resolve()
        if not str(full).startswith(str(base.resolve())):
            raise ValueError("Invalid path")
        if not full.exists() or not full.is_file():
            raise ValueError("File not found")
        return full

    def _preview_csv(path: Path, head: int) -> Dict[str, Any]:
        rows: List[List[Any]] = []
        # 优先UTF-8，失败可考虑回退（此处简单实现）
        with open(path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                rows.append(row)
                if i + 1 >= head:
                    break
        return {"type": "csv", "rows": rows, "columns": len(rows[0]) if rows else 0}

    def _preview_xlsx(path: Path, sheet: Optional[str], head: int) -> Dict[str, Any]:
        try:
            import openpyxl
        except Exception as e:
            raise ValueError(f"openpyxl not available: {e}")
        # 快速校验是否为ZIP容器，非ZIP可能是.xls或被重命名
        if not zipfile.is_zipfile(str(path)):
            # 尝试按.xls解析
            try:
                return _preview_xls(path, head)
            except Exception:
                # 回退为文本/CSV嗅探
                try:
                    return _preview_text(path, head)
                except Exception as e:
                    raise ValueError(f"Not a valid XLSX (zip) file and fallback failed: {e}")
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        try:
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
            rows: List[List[Any]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(["" if (cell is None) else cell for cell in row])
                if i + 1 >= head:
                    break
            return {"type": "xlsx", "sheet": ws.title, "rows": rows, "columns": max((len(r) for r in rows), default=0)}
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def _preview_xls(path: Path, head: int) -> Dict[str, Any]:
        try:
            import xlrd  # 支持旧版 .xls
        except Exception as e:
            raise ValueError(f"xlrd not available for .xls: {e}")
        book = xlrd.open_workbook(str(path))
        sheet = book.sheet_by_index(0)
        nrows = min(head, sheet.nrows)
        rows: List[List[Any]] = []
        for r in range(nrows):
            vals = []
            for c in range(sheet.ncols):
                vals.append(sheet.cell_value(r, c))
            rows.append(vals)
        return {"type": "xls", "sheet": sheet.name, "rows": rows, "columns": sheet.ncols}

    def _preview_text(path: Path, head: int) -> Dict[str, Any]:
        encodings = ["utf-8", "utf-16", "gb18030", "latin-1"]
        last_err: Optional[Exception] = None
        for enc in encodings:
            try:
                with open(path, 'r', encoding=enc, errors='strict') as f:
                    lines: List[str] = []
                    for i, line in enumerate(f):
                        # 去掉末尾换行
                        lines.append(line.rstrip("\r\n"))
                        if i + 1 >= head:
                            break
                return {"type": "text", "encoding": enc, "lines": lines}
            except Exception as e:
                last_err = e
                continue
        raise ValueError(f"Failed to decode text file: {last_err}")

    def _preview_docx(path: Path, head: int) -> Dict[str, Any]:
        try:
            import docx  # python-docx
        except Exception as e:
            raise ValueError(f"python-docx not available: {e}")
        doc = docx.Document(str(path))
        lines: List[str] = []
        for p in doc.paragraphs:
            if p.text is not None:
                lines.append(p.text)
            if len(lines) >= head:
                break
        return {"type": "docx", "lines": lines}

    def preview_uploaded_file_impl(url: str, sheet: Optional[str] = None, head: int = 20) -> Dict[str, Any]:
        path = _resolve_upload_path(url)
        ext = path.suffix.lower()
        if head <= 0:
            head = 20
        meta = {
            "filename": path.name,
            "size_bytes": os.path.getsize(path),
            "ext": ext,
        }
        if ext in (".csv", ".tsv"):
            out = _preview_csv(path, head)
        elif ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            out = _preview_xlsx(path, sheet, head)
        elif ext in (".xls",):
            out = _preview_xls(path, head)
        elif ext in (".txt", ".md", ".json", ".log"):
            out = _preview_text(path, head)
        elif ext in (".docx",):
            out = _preview_docx(path, head)
        else:
            raise ValueError("Unsupported file type; supported: .csv/.tsv/.xlsx/.xls/.txt/.md/.json/.log/.docx")
        out.update(meta)
        return {"ok": True, "preview": out}

    preview_uploaded_file_tool = StructuredTool.from_function(
        func=preview_uploaded_file_impl,
        name="preview_uploaded_file",
        description="Preview a user-uploaded file under /uploads. Supported: .csv/.tsv/.xlsx/.xls/.txt/.md/.json/.log/.docx. Input url can be '/uploads/...' or full http(s) URL.",
        args_schema=PreviewUploadedFileArgs,
    )

    return [
        preview_uploaded_file_tool,
    ]


